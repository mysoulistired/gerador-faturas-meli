"""
Gera faturas (uma planilha .xlsx por hub) a partir da Pré-fatura, usando
o layout de "BETIM - BRMG02.xlsx" como template.

Regras de agregação e mapeamento entre abas foram validadas manualmente
contra faturas oficiais (ver Relatorio_Analise_Pre-Fatura.md). Período e
Line Haul não existem na Pré-fatura e precisam ser passados na linha de
comando.

As colunas são localizadas pelo texto do cabeçalho, não por letra/posição
fixa: o layout da Pré-fatura já mudou de um mês para o outro (ex.: a aba
CT-es ganhou/perdeu uma coluna), o que quebraria índices fixos.

Os CT-es de cada fatura são filtrados pelo CNPJ Tomador do hub (via a
aba "CNPJ BASES"), não pelo nome da cidade: uma mesma cidade pode ter
mais de um hub (ex.: SIMÕES FILHO = BRBA02 e BRXBA1), e filtrar só por
cidade juntava os CT-es dos dois hubs. Isso foi confirmado batendo 100%
com duas faturas oficiais (BETIM e SIMÕES FILHO/BRXBA1).
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import openpyxl
from num2words import num2words

INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*]')


def safe_filename(name: str) -> str:
    """Remove caracteres que o Windows não aceita em nome de arquivo (ex.:
    se a cidade tiver uma barra por erro de digitação, isso quebraria
    wb.save() com um OSError difícil de entender)."""
    return INVALID_FILENAME_CHARS.sub("-", name).strip()


def normalize_cnpj(value) -> str:
    """Alguns arquivos guardam o CNPJ como texto com um apóstrofo inicial
    (marca de 'forçar texto' do Excel); remove isso para comparar."""
    return str(value).strip().lstrip("'")


def find_columns(ws, header_row: int, names: list[str]) -> dict[str, int]:
    """Localiza, na `header_row`, a 1a coluna cujo texto bate com cada nome
    em `names`. Levanta erro claro se algum não for encontrado, em vez de
    seguir em frente lendo a coluna errada silenciosamente."""
    found: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=header_row, column=col).value
        if not isinstance(value, str):
            continue
        text = value.strip()
        for name in names:
            if name not in found and text == name:
                found[name] = col
    faltando = [n for n in names if n not in found]
    if faltando:
        raise ValueError(
            f"Aba '{ws.title}': não encontrei a(s) coluna(s) {faltando!r} "
            f"na linha {header_row}. O layout da planilha pode ter mudado."
        )
    return found


PRESTADOR = {
    "razao_social": "MULTICOM LOG TRANSPORTES LTDA",
    "inscr_est": 150140233110,
    "cnpj": "55.939.622/0001-71",
    "municipio": "Ribeirão Preto/SP",
    "endereco": "R CERQUEIRA CESAR, 1625 - CXPST 44",
    "cep": "14.025-120",
}

CLIENTE = {
    "razao_social": "EBAZAR.COM.BR. LTDA",
    "endereco": "AV DAS NACOES UNIDAS, 3000, BONFIM - PARTE A",
    "municipio": "OSASCO",
    "cnpj": "03.007.331/0001-41",
}

MELI_HEADER_ROW = 2
LANC_HEADER_ROW = 6
CTES_HEADER_ROW = 1


def format_id(value) -> str:
    """Alguns arquivos guardam o número da fatura (ou de CT-e) como float
    (5396.0) em vez de inteiro; sem isso o texto final mostra '5396.0'
    em vez de '5396'."""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def format_brl(value: float) -> str:
    s = f"{value:,.2f}"
    s = s.replace(",", "§").replace(".", ",").replace("§", ".")
    return f"BRL {s}"


def valor_por_extenso(value: float) -> str:
    reais = int(value)
    centavos = round((value - reais) * 100)
    texto = f"{num2words(reais, lang='pt_BR')} reais e {num2words(centavos, lang='pt_BR')} centavos"
    return texto[0].upper() + texto[1:]


def load_meli_maps(ws):
    """Retorna (hub -> cidade, hub -> [(veiculo, nome_rota), ...] em ordem de 1a aparição)."""
    cols = find_columns(ws, MELI_HEADER_ROW, ["Veículo", "Nome da rota", "Origem", "Cidade I"])
    hub_to_city: dict[str, str] = {}
    hub_to_routes: dict[str, dict[tuple, None]] = {}
    for r in range(MELI_HEADER_ROW + 1, ws.max_row + 1):
        hub = ws.cell(row=r, column=cols["Origem"]).value
        cidade = ws.cell(row=r, column=cols["Cidade I"]).value
        veiculo = ws.cell(row=r, column=cols["Veículo"]).value
        rota = ws.cell(row=r, column=cols["Nome da rota"]).value
        if not hub:
            continue
        hub = str(hub).strip()
        if cidade and hub not in hub_to_city:
            hub_to_city[hub] = str(cidade).strip()
        if veiculo and rota:
            hub_to_routes.setdefault(hub, {})[(str(veiculo).strip(), str(rota).strip())] = None
    return hub_to_city, {h: list(d.keys()) for h, d in hub_to_routes.items()}


def load_lancamento_groups(ws):
    """Agrupa a aba 4. Lançamento por hub, só linhas Doc.=='Fatura'.

    'Frete Receita', 'GRIS', 'ICMS/ISS' e 'Total' aparecem duas vezes no
    cabeçalho (bloco Mercado Livre e bloco Lösung); find_columns pega a
    1a ocorrência de cada, que é o bloco Mercado Livre (ver Achado 4 do
    Relatorio_Analise_Pre-Fatura.md sobre por que esse é o bloco certo).
    """
    cols = find_columns(ws, LANC_HEADER_ROW,
                         ["Origem", "Frete Receita", "GRIS", "ICMS/ISS", "Total", "Doc.", "Número"])
    groups: dict[str, dict] = {}
    for r in range(LANC_HEADER_ROW + 1, ws.max_row + 1):
        hub = ws.cell(row=r, column=cols["Origem"]).value
        doc = ws.cell(row=r, column=cols["Doc."]).value
        if not hub or doc != "Fatura":
            continue
        hub = str(hub).strip()
        g = groups.setdefault(hub, {"subtotal": 0.0, "icms": 0.0, "total": 0.0, "numeros": set(), "linhas": 0})
        frete = ws.cell(row=r, column=cols["Frete Receita"]).value or 0
        gris = ws.cell(row=r, column=cols["GRIS"]).value or 0
        icms = ws.cell(row=r, column=cols["ICMS/ISS"]).value or 0
        total = ws.cell(row=r, column=cols["Total"]).value or 0
        numero = ws.cell(row=r, column=cols["Número"]).value
        g["subtotal"] += frete + gris
        g["icms"] += icms
        g["total"] += total
        g["linhas"] += 1
        if numero is not None:
            g["numeros"].add(numero)
    return groups


def load_hub_to_cnpj(ws) -> dict[str, str]:
    """Aba 'CNPJ BASES': sem cabeçalho, coluna A = CNPJ Tomador, coluna C =
    código do hub. É o mapeamento oficial hub -> CNPJ (único por hub, ao
    contrário do nome da cidade, que pode ter mais de um hub)."""
    hub_to_cnpj: dict[str, str] = {}
    for r in range(1, ws.max_row + 1):
        hub = ws.cell(row=r, column=3).value
        cnpj = ws.cell(row=r, column=1).value
        if hub and cnpj:
            hub_to_cnpj[str(hub).strip()] = normalize_cnpj(cnpj)
    return hub_to_cnpj


def load_ctes_by_cnpj(ws) -> dict[str, list[int]]:
    """Indexado pelo CNPJ Tomador (ver load_hub_to_cnpj para o porquê de não
    usar a cidade)."""
    cols = find_columns(ws, CTES_HEADER_ROW, ["Numero", "CNPJ Tomador"])
    by_cnpj: dict[str, list[int]] = {}
    for r in range(CTES_HEADER_ROW + 1, ws.max_row + 1):
        cnpj = ws.cell(row=r, column=cols["CNPJ Tomador"]).value
        numero = ws.cell(row=r, column=cols["Numero"]).value
        if cnpj and numero is not None:
            by_cnpj.setdefault(normalize_cnpj(cnpj), []).append(numero)
    for cnpj in by_cnpj:
        by_cnpj[cnpj].sort()
    return by_cnpj


def build_fatura(template_path: Path, out_path: Path, *, hub: str, cidade: str,
                  emissao, periodo: str, line_haul: str,
                  group: dict, ctes: list[int], rotas: list[tuple]):
    wb = openpyxl.load_workbook(template_path)
    ws = wb.worksheets[0]

    ws["C4"] = PRESTADOR["razao_social"]
    ws["F4"] = PRESTADOR["inscr_est"]
    ws["C5"] = PRESTADOR["cnpj"]
    ws["F5"] = PRESTADOR["municipio"]
    ws["C6"] = PRESTADOR["endereco"]
    ws["F6"] = PRESTADOR["cep"]

    ws["C8"] = CLIENTE["razao_social"]
    ws["C9"] = CLIENTE["endereco"]
    ws["C10"] = CLIENTE["municipio"]
    ws["C11"] = CLIENTE["cnpj"]

    ws["C14"] = emissao  # B14 mantém a fórmula '=90+C14' do template
    numeros_fatura = sorted(format_id(n) for n in group["numeros"])
    if not numeros_fatura:
        print(f"  [!] hub {hub}: nenhum número de fatura preenchido na coluna X, campo ficará vazio.")
    elif len(numeros_fatura) > 1:
        print(f"  [!] hub {hub}: número de fatura não é único na coluna X: {numeros_fatura!r}, "
              f"usando o primeiro.")
    ws["D14"] = numeros_fatura[0] if numeros_fatura else ""
    total = round(group["total"], 2)
    ws["F14"] = format_brl(total)

    ws["B17"] = f"Transporte CTE - {periodo} - Line Haul N. {line_haul}"
    ws["B20"] = valor_por_extenso(total)

    ctes_str = " / ".join(format_id(n) for n in ctes)
    rotas_str = "\n".join(f"{veic.upper()} - {rota}" for veic, rota in rotas)
    ws["B22"] = (
        "Pela sua PRESTAÇÃO DE SERVIÇO conforme Dacte(s) abaixo. Na falta de pagamento "
        "no vencimento, serão cobrados juros de mora.\n\n"
        f"CTE: {ctes_str}\n\n"
        f"{rotas_str}"
    )

    ws["F24"] = round(group["subtotal"], 2)
    ws["F25"] = None
    ws["F26"] = round(group["icms"], 2)
    # F27 mantém a fórmula '=SUM(F24:F26)' do template

    wb.save(out_path)


def load_sheet(wb, name: str):
    try:
        return wb[name]
    except KeyError:
        raise ValueError(
            f"A planilha não tem uma aba chamada '{name}'. Abas encontradas: "
            f"{wb.sheetnames!r}. O layout da Pré-fatura pode ter mudado."
        ) from None


def run_generation(*, prefatura: Path, template: Path, out_dir: Path, emissao,
                    periodo: str, linehaul: str, hub: str | None = None, log=print):
    """Lógica compartilhada pela CLI e pela GUI. `log` recebe cada linha de
    progresso (por padrão, print; a GUI passa sua própria função de log)."""
    log(f"Carregando {prefatura} ...")
    wb = openpyxl.load_workbook(prefatura, data_only=True)
    hub_to_city, hub_to_routes = load_meli_maps(load_sheet(wb, "1. MELI"))
    groups = load_lancamento_groups(load_sheet(wb, "4. Lançamento"))
    hub_to_cnpj = load_hub_to_cnpj(load_sheet(wb, "CNPJ BASES"))
    ctes_by_cnpj = load_ctes_by_cnpj(load_sheet(wb, "CT-es"))

    out_dir.mkdir(parents=True, exist_ok=True)

    hubs = [hub] if hub else sorted(groups.keys())
    falhas = 0
    for h in hubs:
        group = groups.get(h)
        if not group:
            log(f"  [!] hub {h}: nenhuma linha 'Fatura' encontrada na aba 4. Lançamento, pulando.")
            continue
        cidade = hub_to_city.get(h)
        if not cidade:
            log(f"  [!] hub {h}: cidade não encontrada na aba 1. MELI, pulando.")
            continue
        cnpj = hub_to_cnpj.get(h)
        if not cnpj:
            log(f"  [!] hub {h}: não encontrado na aba CNPJ BASES, não dá pra filtrar os CT-es com segurança.")
            ctes = []
        else:
            ctes = ctes_by_cnpj.get(cnpj, [])
            if not ctes:
                log(f"  [!] hub {h} ({cidade}): nenhum CT-e encontrado na aba CT-es para o CNPJ {cnpj}.")
        rotas = hub_to_routes.get(h, [])

        out_path = out_dir / f"{safe_filename(cidade)} - {safe_filename(h)}.xlsx"
        try:
            build_fatura(
                template, out_path,
                hub=h, cidade=cidade, emissao=emissao,
                periodo=periodo, line_haul=linehaul,
                group=group, ctes=ctes, rotas=rotas,
            )
        except Exception as exc:
            falhas += 1
            log(f"  [ERRO] hub {h} ({cidade}): falha ao gerar a fatura, pulando. Detalhe: {exc}")
            continue
        log(f"  OK  {out_path.name}  "
            f"(subtotal={group['subtotal']:.2f} icms={group['icms']:.2f} total={group['total']:.2f} "
            f"linhas={group['linhas']} ctes={len(ctes)})")

    if falhas:
        log(f"Concluído com {falhas} hub(s) que falharam (veja os [ERRO] acima).")
    else:
        log("Concluído.")


def main():
    import datetime

    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--prefatura", required=True, type=Path,
                     help="Caminho da Pré-fatura (.xlsx)")
    ap.add_argument("--template", required=True, type=Path,
                     help="Caminho do template de fatura (ex.: BETIM - BRMG02.xlsx)")
    ap.add_argument("--out-dir", required=True, type=Path,
                     help="Pasta onde salvar as faturas geradas")
    ap.add_argument("--emissao", required=True,
                     help="Data de emissão (YYYY-MM-DD), usada em todas as faturas geradas")
    ap.add_argument("--periodo", required=True,
                     help='Código de período p/ Observação (ex.: "202608Q1")')
    ap.add_argument("--linehaul", required=True,
                     help='Número "Line Haul" p/ Observação (ex.: "6745030")')
    ap.add_argument("--hub", default=None,
                     help="Gerar só um hub específico (ex.: BRMG02). Default: todos.")
    args = ap.parse_args()

    run_generation(
        prefatura=args.prefatura, template=args.template, out_dir=args.out_dir,
        emissao=datetime.date.fromisoformat(args.emissao),
        periodo=args.periodo, linehaul=args.linehaul, hub=args.hub,
    )


if __name__ == "__main__":
    main()
